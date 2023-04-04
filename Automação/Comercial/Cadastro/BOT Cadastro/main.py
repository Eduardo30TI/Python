from openpyxl import load_workbook
import os
from glob import glob
from Acesso import Login
from Query import Query
from Email import Email
from RemoverArquivo import Remover
import win32com.client as win32
from datetime import datetime
from WebTransfer import Web
import shutil
from zipfile import ZipFile

s=Login()

sql=Query(s.usuario,s.senha,s.database,s.server)

querys={

    'Fornecedores':

    """
    
    SELECT * FROM netfeira.vw_fornecedores
    WHERE [Nome Fantasia] LIKE '%NETFEIRA%' AND [Plano de Conta]='FORNECEDORES'
    
    """,

    'Produtos':

    """
    
    SELECT * FROM netfeira.vw_prod_dados
    
    """

}

subjects={'cadastro joanin':'JOANIN','dados dos produtos':'Produtos','cadastro chocolândia': 'CHOCOLANDIA','cadastro bergamini':'BERGAMINI','cadastro rossi':'ROSSI'}

def Main():

    outlook=win32.Dispatch('Outlook.Application').GetNameSpace('MAPI')

    inbox=outlook.GetDefaultFolder(6)

    emails=[l for l in inbox.items if str(l.subject).lower() in subjects.keys()]

    col_leach='códigos'

    for e in emails:

        func=subjects[str(e.subject).lower()]

        msg='Bom dia' if datetime.now().hour<12 else 'Boa tarde'

        email_to=[e.sender]

        email_cc=[l for l in str(e.cc).split(';') if str(l).find(inbox.Parent.Name)<0]

        anexo=[]

        assunto=str(e.subject).capitalize()

        try:

            elemento=[l for l in str(e.body).strip().split('\n') if str(l).lower().find(col_leach)>=0][-1]

            elemento=[int(l) for l in elemento.split()[-1].split(',') if str(l).isnumeric()]
            
            temp_path=globals().get(func)(elemento)

            anexo.append(temp_path)

            temp_path=os.path.join(os.getcwd(),'*.zip')

            zips=glob(temp_path)

            if len(zips)>0:

                for zip in zips:

                    w=Web(zip)

                    link_href=w.WebLink()

                    pass

                pass

            mensagem=f"""
            
            <p>{msg};</p>

            <p>Segue a sua solicitação conforme solicitado.</p>

            <P>Por favor não responder mensagem automática</P>

            <p>Atenciosamente</p>

            <p>BOT TI</p>
            
            """ if len(zips)<=0 else f"""
            
            <p>{msg};</p>

            <p>Segue a sua solicitação conforme solicitado. E abaixo segue o link das fotos: </p>

            <p>Link com as fotos: <a href="{link_href}">{link_href}</a></p>

            <P>Por favor não responder mensagem automática</P>

            <p>Atenciosamente</p>

            <p>BOT TI</p>
            
            """ 

            temp_dict={'To':email_to,'CC':email_cc,'Anexo':anexo}

            Email.EnviarEmail(corpo=mensagem,assunto=assunto,kwargs=temp_dict)
            
            pass

        except:

            mensagem=f"""

            <p>{msg};</p>

            <p>Identificamos que na sua solicitação não contém dados para serem analisados. Por favor preencher conforme exemplo abaixo: </p>

            <P><strong>Ex: inserir no corpo do e-mail "códigos:" e os códigos com virgula para que possamos atender a solicitação caso haja mais de um.</strong></P>

            <P>Por favor não responder mensagem automática</P>

            <p>Atenciosamente</p>

            <p>BOT TI</p>
            
            """

            temp_dict={'To':email_to,'CC':email_cc,'Anexo':[]}

            Email.EnviarEmail(corpo=mensagem,assunto=assunto,kwargs=temp_dict)            

            pass

        for m in inbox.Folders:

            try:
                
                inbox.Folders.Add(func)

                pass

            except:

                continue

            pass

        path_move=inbox.Folders[func]

        e.move(path_move)

        for opc in ['.xlsx','.zip']:

            Remover.RemoverArquivo(opc)

            pass

        pass

    pass

def JOANIN(produtos:list):

    name_arq='Cadastro JOANIN.xlsx'

    df=sql.CriarTabela(kwargs=querys)

    col_leach='joanin'

    path_base=os.path.join(os.getcwd(),'Planilhas','*.xlsx')

    df['Produtos']=df['Produtos'].loc[df['Produtos']['SKU'].isin(produtos)]
    
    for arq in glob(path_base):

        arq_name=os.path.basename(arq)

        if str(arq_name).lower().find(col_leach)<0:

            continue
        
        temp_path=arq

        pass

    sheet=load_workbook(temp_path)
    sheet.active

    sheet_names=sheet.sheetnames[-1]

    range=sheet[sheet_names]

    for c in df['Produtos']['Fotos'].tolist():

        try:

            nome_arq=os.path.basename(c)

            origem=c

            destino=os.path.join(os.getcwd(),'Fotos',nome_arq)

            if not os.path.exists(os.path.dirname(destino)):
            
                os.makedirs(os.path.dirname(destino))

                pass

            shutil.copy(origem,destino)

            pass

        except:

            continue

        pass

    if os.path.exists(os.path.dirname(destino)):

        dir_name=os.path.basename(os.path.dirname(destino))

        temp_path=os.path.dirname(destino)

        shutil.make_archive(dir_name,'zip',temp_path)

        shutil.rmtree(os.path.dirname(destino))

        pass
        
    for i,c in enumerate(df['Produtos']['SKU'].tolist()):
        
        if i==0:

            name_sheet=str(c)
            range.title=name_sheet

            pass

        else:

            name_sheet=str(c)
            range=sheet.copy_worksheet(range)
            range.title=name_sheet
            
            pass

        temp_dict={'C5':'Razão Social','C6':'CNPJ Formatado','F6':'Inscrição Formatada','C7':'Endereço','C8':'Bairro','C9':'Município','C10':'Complemento','F7':'Número','F8':'CEP','F9':'UF'}
        
        #preencher dados do fornecedor
        sheet_names=sheet.sheetnames[-1]

        for key,value in temp_dict.items():
            
            range[key]=str(df['Fornecedores'][value].max())

            #break

            pass

        temp_dict={'C16':'Produto','C17':'EAN13', 'F17':'Fabricante','C18':'Peso Líquido KG', 'F18':'Peso Bruto KG', 'C19':'Comprimento STK', 'C20':'Largura STK', 'C21': 'Altura STK', 'C23': 'SKU','C26':'Unid. VDA', 'C27':'Qtde Caixa', 'C28': 'DUN14', 'C29': 'Peso Caixa Bruto','C30': 'Peso Caixa Líquido','C31': 'Comprimento CMP','C32': 'Largura CMP', 'C33':'Altura CMP', 'C47':'NCM', 'C48':'CEST'}
        
        #preencher dados dos produtos
        for key,value in temp_dict.items():
            
            range[key]=str(df['Produtos'].loc[df['Produtos']['SKU']==c,value].max())
            #break

            pass

        pass

    sheet.save(name_arq)

    temp_path=os.path.join(os.getcwd(),name_arq)
    
    return temp_path

    pass

def Produtos(produtos:list):

    df=sql.CriarTabela(kwargs=querys)

    df['Produtos']=df['Produtos'].loc[df['Produtos']['SKU'].isin(produtos)]

    df['Produtos'].to_excel('Dados cadastrais.xlsx',index=False)

    temp_path=os.path.join(os.getcwd(),'*.xlsx')

    arquivos=glob(temp_path)

    for c in df['Produtos']['Fotos'].tolist():

        nome_arq=os.path.basename(c)

        origem=c

        destino=os.path.join(os.getcwd(),'Fotos',nome_arq)

        if not os.path.exists(os.path.dirname(destino)):
        
            os.makedirs(os.path.dirname(destino))

            pass

        shutil.copy(origem,destino)

        pass

    if os.path.exists(os.path.dirname(destino)):

        dir_name=os.path.basename(os.path.dirname(destino))

        temp_path=os.path.dirname(destino)

        shutil.make_archive(dir_name,'zip',temp_path)

        shutil.rmtree(os.path.dirname(destino))

        pass    

    return arquivos[-1]

    pass

def CHOCOLANDIA(produtos:list):
    
    name_arq='Cadastro CHOCOLÂNDIA.xlsx'

    df=sql.CriarTabela(kwargs=querys)

    col_leach='chocolândia'

    path_base=os.path.join(os.getcwd(),'Planilhas','*.xlsx')

    df['Produtos']=df['Produtos'].loc[df['Produtos']['SKU'].isin(produtos)].reset_index()
    
    for arq in glob(path_base):

        arq_name=os.path.basename(arq)

        if str(arq_name).lower().find(col_leach)<0:

            continue
        
        temp_path=arq

        pass

    sheet=load_workbook(temp_path)
    sheet.active

    sheet_names=sheet.sheetnames[-1]

    range=sheet[sheet_names]

    for c in df['Produtos']['Fotos'].tolist():
        
        try:

            nome_arq=os.path.basename(c)

            origem=c

            destino=os.path.join(os.getcwd(),'Fotos',nome_arq)

            if not os.path.exists(os.path.dirname(destino)):
            
                os.makedirs(os.path.dirname(destino))

                pass

            shutil.copy(origem,destino)

            pass

        except:

            continue

        pass


    if os.path.exists(os.path.dirname(destino)):

        dir_name=os.path.basename(os.path.dirname(destino))

        temp_path=os.path.dirname(destino)

        shutil.make_archive(dir_name,'zip',temp_path)

        shutil.rmtree(os.path.dirname(destino))

        pass

    lista = df['Produtos'].index.tolist()
    for i,c in enumerate(['SKU','Produto']):
        indice=3
        for l in lista:
            val = df['Produtos'].loc[l,c]
            range.cell(row=indice,column=i+1,value=val)
            indice+=1

            pass

        pass

    sheet.save(name_arq)

    temp_path=os.path.join(os.getcwd(),name_arq)

    return temp_path    

    pass

def BERGAMINI(produtos: list):

    name_arq='Cadastro BERGAMINI.xlsx'

    df=sql.CriarTabela(kwargs=querys)

    col_leach='bergamini'

    path_base=os.path.join(os.getcwd(),'Planilhas','*.xlsx')

    df['Produtos']=df['Produtos'].loc[df['Produtos']['SKU'].isin(produtos)]
    
    for arq in glob(path_base):

        arq_name=os.path.basename(arq)

        if str(arq_name).lower().find(col_leach)<0:

            continue
        
        temp_path=arq

        pass

    sheet=load_workbook(temp_path)
    sheet.active

    sheet_names=sheet.sheetnames[-1]

    range=sheet[sheet_names]

    for c in df['Produtos']['Fotos'].tolist():

        try:

            nome_arq=os.path.basename(c)

            origem=c

            destino=os.path.join(os.getcwd(),'Fotos',nome_arq)

            if not os.path.exists(os.path.dirname(destino)):
            
                os.makedirs(os.path.dirname(destino))

                pass

            shutil.copy(origem,destino)

            pass

        except:

            continue

        pass

    if os.path.exists(os.path.dirname(destino)):

        dir_name=os.path.basename(os.path.dirname(destino))

        temp_path=os.path.dirname(destino)

        shutil.make_archive(dir_name,'zip',temp_path)

        shutil.rmtree(os.path.dirname(destino))

        pass

    #preencher planilha
    for i,c in enumerate(df['Produtos']['SKU'].tolist()):
        
        if i==0:

            name_sheet=str(c)
            range.title=name_sheet

            pass

        else:

            name_sheet=str(c)
            range=sheet.copy_worksheet(range)
            range.title=name_sheet
            
            pass

        temp_dict={'AA12':'Razão Social'}
        
        #preencher dados do fornecedor
        sheet_names=sheet.sheetnames[-1]

        for key,value in temp_dict.items():
            
            range[key]=str(df['Fornecedores'][value].max())

            #break

            pass

        temp_dict={'K6':'Produto','I8':'SKU','J10':'Fabricante','E12':'Qtde Caixa',4:'EAN13',10:'DUN14',8:'NCM'}

        col_leach={'EAN13':9,'DUN14':26,'NCM':31}
        
        #preencher dados dos produtos
        for key,value in temp_dict.items():

            if key in [4,10,8]:

                val=str(df['Produtos'].loc[df['Produtos']['SKU']==c,value].max())

                if key==8:

                    val=val.strip().replace('.','')

                    pass

                lista=[l for l in val]

                indice=col_leach[value]

                for i,j in enumerate(lista):
                                        
                    range.cell(row=key,column=indice+i,value=str(j))

                    pass

                pass

            else:
                                    
                range[key]=str(df['Produtos'].loc[df['Produtos']['SKU']==c,value].max())

                pass
            #break

            pass
        
        pass


    sheet.save(name_arq)

    temp_path=os.path.join(os.getcwd(),name_arq)
    
    return temp_path

    pass

def ROSSI(produtos: list):

    name_arq='Cadastro ROSSI.xlsx'

    df=sql.CriarTabela(kwargs=querys)

    col_leach='rossi'

    path_base=os.path.join(os.getcwd(),'Planilhas','*.xlsx')

    df['Produtos']=df['Produtos'].loc[df['Produtos']['SKU'].isin(produtos)].reset_index()
    
    for arq in glob(path_base):

        arq_name=os.path.basename(arq)

        if str(arq_name).lower().find(col_leach)<0:

            continue
        
        temp_path=arq

        pass

    sheet=load_workbook(temp_path)
    sheet.active

    sheet_names=sheet.sheetnames[-1]

    range=sheet[sheet_names]

    for c in df['Produtos']['Fotos'].tolist():

        try:

            nome_arq=os.path.basename(c)

            origem=c

            destino=os.path.join(os.getcwd(),'Fotos',nome_arq)

            if not os.path.exists(os.path.dirname(destino)):
            
                os.makedirs(os.path.dirname(destino))

                pass

            shutil.copy(origem,destino)

            pass

        except:

            continue

        pass

    if os.path.exists(os.path.dirname(destino)):

        dir_name=os.path.basename(os.path.dirname(destino))

        temp_path=os.path.dirname(destino)

        shutil.make_archive(dir_name,'zip',temp_path)

        shutil.rmtree(os.path.dirname(destino))

        pass

    col={'Altura CMP':22,'Largura CMP':23,'Comprimento CMP':24}

    #preencher planilha
    lista = df['Produtos'].index.tolist()
    for i,c in enumerate(['SKU','Produto','Fabricante','Unid. VDA','Qtde Caixa','EAN13','DUN14','NCM','CST VDA','CEST','Altura CMP','Largura CMP','Comprimento CMP']):
        indice=6
        for l in lista:

            val = df['Produtos'].loc[l,c] if not c in ['NCM','CEST'] else str(df['Produtos'].loc[l,c]).replace('.','')
            
            if c in col.keys():

                range.cell(row=indice,column=col[c],value=val*100)
                indice+=1

                pass

            else:
                
                range.cell(row=indice,column=i+1,value=val)
                indice+=1

                pass


            pass

        pass
    
    sheet.save(name_arq)

    temp_path=os.path.join(os.getcwd(),name_arq)

    return temp_path

    pass

if __name__=='__main__':
        
    Main()

    pass